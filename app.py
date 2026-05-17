from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import datetime
import random
import re
import os
import io

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, '2026-2027 Attending Call - Vacay_Away Dates.xlsx')

# ─── Name / status lookups ────────────────────────────────────────────────────

NAME_MAP = {
    "Abou-Jaoude":  "Michelle Abou-Jaoude",
    "Abou-jaoude":  "Michelle Abou-Jaoude",
    "Agarwal":      "Manokamna Agarwal",
    "Bansal":       "Aditya Bansal",
    "Bastos":       "Ana Bastos de Carvalho",
    "Blackburn":    "Peter Blackburn",
    "Bradley":      "Melanie Bradley",
    "Capoor":       "Seema Capoor",
    "Clevenger":    "Leanne Clevenger",
    "Franklin, J":  "John Franklin",
    "Franklin, L":  "Lucy Franklin",
    "J. Franklin":  "John Franklin",
    "L. Franklin":  "Lucy Franklin",
    "L. Frankin":   "Lucy Franklin",
    "Fraser":       "Claire Fraser",
    "Katz":         "Douglas Katz",
    "Marzolf":      "Amanda Marzolf",
    "Moore":        "Daniel Moore",
    "Oberst":       "Garrett Oberst",
    "Rust":         "Christine Rust",
    "Tannen":       "Bradford Tannen",
    "Wadley":       "Sean Wadley",
    "Willen":       "Christi Willen",
    # historical (holiday counter only)
    "Pearson":      "Andrew Pearson",
    "Conklin":      "John Conklin",
    "Sanders":      "Sheila Sanders",
    "Young":        "Lauren Young",
    "Higgins":      "Higgins",
    # 2026-2027 Google Form variations
    "Danny Moore":          "Daniel Moore",
    "Manokamna Agarwal":    "Manokamna Agarwal",
    "Bradford L Tannen":    "Bradford Tannen",
}

STATUS_DEFAULTS = {
    "Michelle Abou-Jaoude":    "full time",
    "Manokamna Agarwal":      "full time",
    "Aditya Bansal":           "full time",
    "Ana Bastos de Carvalho":  "full time",
    "Peter Blackburn":         "full time",
    "Melanie Bradley":         "full time",
    "Seema Capoor":            "full time",
    "Leanne Clevenger":        "full time",
    "John Franklin":           "full time",
    "Lucy Franklin":           "part time",
    "Claire Fraser":           "full time",
    "Douglas Katz":            "full time",
    "Amanda Marzolf":          "part time 1",
    "Daniel Moore":            "full time",
    "Garrett Oberst":          "full time",
    "Christine Rust":          "full time",
    "Bradford Tannen":         "full time",
    "Sean Wadley":             "full time",
    "Christi Willen":          "part time",
}

DEFAULT_HOLIDAYS = {
    "labor_day":        "2026-09-07",
    "thanksgiving":     "2026-11-26",
    "black_friday":     "2026-11-27",
    "christmas":        "2026-12-25",
    "new_years_eve":    "2026-12-31",
    "new_years":        "2027-01-01",
    "mlk_day":          "2027-01-19",
    "memorial_day":     "2027-05-31",
    "juneteenth":       "2027-06-19",
    "independence_day": "2027-07-05",
}

# ─── Date parsing ─────────────────────────────────────────────────────────────

_MONTHS = {
    'jan':1,'january':1,'feb':2,'february':2,'mar':3,'march':3,
    'apr':4,'april':4,'may':5,'jun':6,'june':6,'jul':7,'july':7,
    'aug':8,'august':8,'sep':9,'sept':9,'september':9,
    'oct':10,'october':10,'nov':11,'november':11,'dec':12,'december':12,
}

def _month(s):
    return _MONTHS.get(s.lower().rstrip('.'), None)

def _infer_year(month):
    return 2026 if month >= 7 else 2027

def _fmt(d):
    return f"{d.month}/{d.day}/{str(d.year)[2:]}"

def _strip_ord(s):
    return re.sub(r'(\d+)(st|nd|rd|th)\b', r'\1', s, flags=re.I)

_WEEKDAY_RE = re.compile(
    r'\b(monday|tuesday|wednesday|weds|thursday|friday|saturday|sunday|'
    r'mon|tue|wed|thu|thur|fri|sat|sun)\b\s*',
    re.I
)

# Words/phrases that appear as trailing labels after a date
_TRAIL_RE = re.compile(
    r'(?<=\d)[\s\-–]+(vacation|conference|conferences|personal|professional|pto|'
    r'aao|aupo|aapos|arvo|abo|asrs|vbs|ascrs|sonoma|floretina|birthday|anniversary|'
    r'reunion|trip|game|foster|holiday|surgery|meeting|sporadically|preferred|'
    r'fall|spring|summer|winter|week|call|out of town|soccer|paternity|possible|'
    r'weekend|fmla|pm)\b.*$',
    re.I
)
# Words that can appear at the start (before the date)
_LEAD_RE = re.compile(
    r'^(vacation|conference|personal|professional|aao|aupo|aapos|arvo|abo|asrs|'
    r'vbs|ascrs|sonoma|floretina|birthday|anniversary|reunion|break|trip|game|foster|'
    r'holiday|surgery|meeting|sporadically|preferred|guatemala|ukgo|uk go|paternity|'
    r'memorial day|labor day|independence day|week of|all of|please note|possible)\s+',
    re.I
)

def _clean(entry):
    entry = entry.strip()
    # Strip ordinal suffixes first so they don't confuse later patterns
    entry = _strip_ord(entry)
    # Strip weekday names (Monday, Friday, etc.) — they carry no scheduling info
    entry = _WEEKDAY_RE.sub('', entry).strip()
    # Remove parenthetical notes
    entry = re.sub(r'\([^)]*\)', '', entry)
    # Remove trailing punctuation and markdown decorators
    entry = re.sub(r'[*#~]+$', '', entry).rstrip(' .;,')
    # Remove "sentence continues" after a period+space+uppercase
    entry = re.sub(r'\.\s+[A-Z].*$', '', entry)
    entry = re.sub(r'[*#~]+$', '', entry).rstrip(' .;,')
    # Handle colon: "date: label" → keep date; "label: date" → keep date
    colon_m = re.match(r'^(.+?)\s*:\s*(.+)$', entry)
    if colon_m:
        before, after = colon_m.group(1).strip(), colon_m.group(2).strip()
        if any(_month(w) for w in re.split(r'\W+', before)):
            entry = before   # "June 15: note" → keep "June 15"
        else:
            entry = after    # "Paternity: July 1-9" → keep "July 1-9"
    # Remove leading label words (run iteratively to strip multi-word prefixes)
    for _ in range(4):
        prev = entry
        entry = _LEAD_RE.sub('', entry)
        if entry == prev:
            break
    # Remove trailing label words (after a digit)
    entry = _TRAIL_RE.sub('', entry)
    # Remove any remaining trailing all-alpha suffix after a digit (e.g. "Oct 17-19 AAO")
    entry = re.sub(r'(?<=\d)[\s\-–]+[A-Za-z][A-Za-z\s]*$', '', entry)
    # Strip "- Description" suffix where description is not a month (e.g. "Apr 3 - Vit Buckle Society")
    def _strip_dash_desc(s):
        def repl(m):
            return '' if not _month(m.group(1)) else m.group(0)
        return re.sub(r'\s*[-–]\s+([A-Za-z]\w*).*$', repl, s, flags=re.I)
    entry = _strip_dash_desc(entry)
    # Fix "DD- YYYY" → "DD YYYY" (stray dash before year)
    entry = re.sub(r'(\d)\s*-\s*(\d{4})', r'\1 \2', entry)
    # Normalize " to " → "-"
    entry = re.sub(r'\s+to\s+', '-', entry, flags=re.I)
    # Strip any leading non-month, non-digit word that slipped through
    for _ in range(3):
        m = re.match(r'^([A-Za-z]+)\s+(.+)', entry, re.I)
        if m and not _month(m.group(1)):
            entry = m.group(2)
        else:
            break
    entry = entry.strip(' -.–,')
    return entry

def _parse_one(raw):
    entry = _clean(raw)
    if not entry:
        return []

    def safe_date(yr, mo, d):
        try:
            return datetime.date(yr, mo, d)
        except ValueError:
            return None

    # MM/DD/YY-MM/DD/YY  or  MM/DD/YYYY-MM/DD/YYYY
    m = re.match(r'^(\d{1,2}/\d{1,2}/\d{2,4})\s*[-–]\s*(\d{1,2}/\d{1,2}/\d{2,4})$', entry)
    if m:
        def pmdy(s):
            p = s.split('/')
            yr = int(p[2]); yr = 2000+yr if yr < 100 else yr
            return safe_date(yr, int(p[0]), int(p[1]))
        d1, d2 = pmdy(m.group(1)), pmdy(m.group(2))
        if d1 and d2:
            return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # MM/DD/YY single
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$', entry)
    if m:
        yr = int(m.group(3)); yr = 2000+yr if yr < 100 else yr
        d = safe_date(yr, int(m.group(1)), int(m.group(2)))
        if d: return [_fmt(d)]

    # Month D-D [YYYY]  (same month)
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})\s*[-–]\s*(\d{1,2})\s*,?\s*(\d{4})?$', entry)
    if m:
        mo = _month(m.group(1))
        if mo:
            yr = int(m.group(4)) if m.group(4) else _infer_year(mo)
            d1 = safe_date(yr, mo, int(m.group(2)))
            d2 = safe_date(yr, mo, int(m.group(3)))
            if d1 and d2 and d1 <= d2: return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # Month D - Month D [YYYY]  (cross-month)
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})\s*[-–]\s*([A-Za-z]+)\s*(\d{1,2})\s*,?\s*(\d{4})?$', entry)
    if m:
        mo1, mo2 = _month(m.group(1)), _month(m.group(3))
        if mo1 and mo2:
            yr = int(m.group(5)) if m.group(5) else _infer_year(mo1)
            yr2 = yr if mo2 >= mo1 else yr+1
            d1 = safe_date(yr, mo1, int(m.group(2)))
            d2 = safe_date(yr2, mo2, int(m.group(4)))
            if d1 and d2: return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # Month D through D [YYYY]
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})\s+through\s+(\d{1,2})\s*,?\s*(\d{4})?$', entry, re.I)
    if m:
        mo = _month(m.group(1))
        if mo:
            yr = int(m.group(4)) if m.group(4) else _infer_year(mo)
            d1 = safe_date(yr, mo, int(m.group(2)))
            d2 = safe_date(yr, mo, int(m.group(3)))
            if d1 and d2: return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # Month D through Month D [YYYY]
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})\s+through\s+([A-Za-z]+)\s+(\d{1,2})\s*,?\s*(\d{4})?$', entry, re.I)
    if m:
        mo1, mo2 = _month(m.group(1)), _month(m.group(3))
        if mo1 and mo2:
            yr = int(m.group(5)) if m.group(5) else _infer_year(mo1)
            yr2 = yr if mo2 >= mo1 else yr+1
            d1 = safe_date(yr, mo1, int(m.group(2)))
            d2 = safe_date(yr2, mo2, int(m.group(4)))
            if d1 and d2: return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # Month D YYYY  (single with year)
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})\s*,?\s*(\d{4})$', entry)
    if m:
        mo = _month(m.group(1))
        if mo:
            d = safe_date(int(m.group(3)), mo, int(m.group(2)))
            if d: return [_fmt(d)]

    # Month D  (single, infer year)
    m = re.match(r'^([A-Za-z]+)\s+(\d{1,2})$', entry)
    if m:
        mo = _month(m.group(1))
        if mo:
            d = safe_date(_infer_year(mo), mo, int(m.group(2)))
            if d: return [_fmt(d)]

    # MM/DD-MM/DD/YY[YY]  (Tannen style: year only on second part)
    m = re.match(r'^(\d{1,2}/\d{1,2})\s*[-–]\s*(\d{1,2}/\d{1,2}/\d{2,4})$', entry)
    if m:
        p2 = m.group(2).split('/')
        yr = int(p2[2]); yr = 2000+yr if yr < 100 else yr
        p1 = m.group(1).split('/')
        d1 = safe_date(yr, int(p1[0]), int(p1[1]))
        d2 = safe_date(yr, int(p2[0]), int(p2[1]))
        if d1 and d2: return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # MM/DD/YY MM/DD/YY  (two dates separated by space, e.g. "12/19/25 1/1/26")
    m = re.match(r'^(\d{1,2}/\d{1,2}/\d{2,4})\s+(\d{1,2}/\d{1,2}/\d{2,4})$', entry)
    if m:
        def pmdy2(s):
            p = s.split('/'); yr = int(p[2]); yr = 2000+yr if yr < 100 else yr
            return safe_date(yr, int(p[0]), int(p[1]))
        d1, d2 = pmdy2(m.group(1)), pmdy2(m.group(2))
        if d1 and d2: return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # MM/DD-MM/DD  (no year, e.g. Moore style "7/25-7/26" or cross-month "12/24-1/3")
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*[-–]\s*(\d{1,2})/(\d{1,2})$', entry)
    if m:
        mo1, d1v, mo2, d2v = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
        yr1 = _infer_year(mo1)
        yr2 = yr1 if mo2 >= mo1 else yr1 + 1
        d1 = safe_date(yr1, mo1, d1v)
        d2 = safe_date(yr2, mo2, d2v)
        if d1 and d2: return [f"{_fmt(d1)}-{_fmt(d2)}"]

    # MM/DD  (single no-year, e.g. "2/7")
    m = re.match(r'^(\d{1,2})/(\d{1,2})$', entry)
    if m:
        mo, dv = int(m.group(1)), int(m.group(2))
        d = safe_date(_infer_year(mo), mo, dv)
        if d: return [_fmt(d)]

    return []

_MDY_RANGE = re.compile(r'\d{1,2}/\d{1,2}/\d{2,4}\s*[-–]\s*\d{1,2}/\d{1,2}/\d{2,4}')
_MDY_SINGLE = re.compile(r'\d{1,2}/\d{1,2}/\d{2,4}')

def parse_dates(text):
    """Return (parsed_list, failed_lines) from free-text field."""
    if not text or str(text).strip() in ('N/A', 'NA', 'None', ''):
        return [], []
    parsed, failed = [], []
    # Normalize semicolons, commas between date entries, and "and" to newlines
    normalized = re.sub(r';', '\n', str(text))
    normalized = re.sub(r'\band\b(?=\s+\d)', '\n', normalized, flags=re.I)
    # Split "Month D-D, Month D-D" style comma-separated lists
    # Only split on ", " when the part after the comma looks like a date start (month word or digit)
    normalized = re.sub(r',\s+(?=[A-Za-z]{3,}\s+\d|\d{1,2}[/\-])', '\n', normalized)
    for line in normalized.split('\n'):
        line = line.strip()
        if not line or not re.search(r'\d', line):
            continue
        results = _parse_one(line)
        if results:
            parsed.extend(results)
        else:
            # Fallback: extract embedded MM/DD/YY ranges or single dates from long text
            extracted = []
            for rng in _MDY_RANGE.findall(line):
                extracted.extend(_parse_one(rng))
            if not extracted:
                seen = set()
                for single in _MDY_SINGLE.findall(line):
                    if single not in seen:
                        seen.add(single)
                        extracted.extend(_parse_one(single))
            if extracted:
                parsed.extend(extracted)
            elif re.search(r'\d', line):
                failed.append(line)
    return parsed, failed

# ─── Excel parsing ────────────────────────────────────────────────────────────

def _lookup(raw):
    raw = str(raw).strip().strip("'\"")
    return NAME_MAP.get(raw, NAME_MAP.get(raw.title(), raw))

def _parse_attending_row(raw_name, prefs_raw, abs_raw):
    prefs_raw = '' if prefs_raw in ('N/A', 'NA') else prefs_raw
    abs_raw   = '' if abs_raw   in ('N/A', 'NA') else abs_raw
    full_name = _lookup(raw_name)
    status    = STATUS_DEFAULTS.get(full_name, 'full time')
    parsed_prefs, warn_prefs = parse_dates(prefs_raw)
    parsed_abs,   warn_abs   = parse_dates(abs_raw)
    return {
        'name':             full_name,
        'status':           status,
        'preferences':      parsed_prefs,
        'absences':         parsed_abs,
        'raw_preferences':  prefs_raw,
        'raw_absences':     abs_raw,
        'warn_preferences': warn_prefs,
        'warn_absences':    warn_abs,
    }

def parse_excel(path_or_file):
    wb = openpyxl.load_workbook(path_or_file)
    ws1 = wb.worksheets[0]

    # Detect format: new Google Form export has "Timestamp" in A1
    is_new_format = str(ws1.cell(1, 1).value or '').strip().lower() == 'timestamp'

    if is_new_format:
        # New format (2026-2027): single sheet, columns:
        #   1=Timestamp, 2=Name, 3=Vacation/absences, 4=Not-able dates, 5=Preferences
        attendings = []
        for row in range(2, ws1.max_row + 1):
            raw_name = ws1.cell(row, 2).value
            if not raw_name:
                continue
            raw_name = str(raw_name).strip().strip("'\"")

            col3 = str(ws1.cell(row, 3).value or '').strip()
            col4 = str(ws1.cell(row, 4).value or '').strip()
            # Combine general absences (col3) and "not able" dates (col4)
            abs_parts = [p for p in [col3, col4] if p and p not in ('N/A', 'NA', 'na', 'none', 'None')]
            abs_raw   = '\n'.join(abs_parts)

            prefs_raw = str(ws1.cell(row, 5).value or '').strip()

            attendings.append(_parse_attending_row(raw_name, prefs_raw, abs_raw))

        # Read Holiday Counter sheet if present
        hc_sheet = next(
            (wb.worksheets[i] for i, s in enumerate(wb.sheetnames)
             if 'holiday' in s.lower() and 'counter' in s.lower()),
            None
        )
        if hc_sheet:
            # Collect data rows (col 1 is a rank number)
            data_rows = []
            for r in range(1, hc_sheet.max_row + 1):
                v = hc_sheet.cell(r, 1).value
                if v is not None and str(v).strip().isdigit():
                    data_rows.append(r)

            def _names_from_row(row_num):
                names = set()
                for c in range(2, hc_sheet.max_column + 1):
                    v = hc_sheet.cell(row_num, c).value
                    if v:
                        s = str(v).strip().strip("'\"")
                        if s and s.lower() not in ('none', 'n/a', 'na'):
                            names.add(_lookup(s))
                return names

            last_year  = _names_from_row(data_rows[-1])  if len(data_rows) >= 1 else set()
            two_yr_ago = _names_from_row(data_rows[-2])  if len(data_rows) >= 2 else set()
            holiday_data = {
                'last_year':          sorted(last_year),
                'two_years_ago':      sorted(two_yr_ago),
                'holidays_two_years': sorted(last_year | two_yr_ago),
            }
        else:
            holiday_data = {
                'last_year':          [],
                'two_years_ago':      [],
                'holidays_two_years': [],
            }
        return attendings, holiday_data

    else:
        # Old format (2025-2026): two sheets, rows start at 4
        ws2 = wb.worksheets[1]

        attendings = []
        for row in range(4, ws1.max_row + 1):
            raw_name = ws1.cell(row, 1).value
            if not raw_name:
                continue
            raw_name = str(raw_name).strip().strip("'\"")
            if re.search(r'aged.?out', raw_name, re.I):
                break

            prefs_raw = str(ws1.cell(row, 2).value or '').strip().strip("'\"")
            abs_raw   = str(ws1.cell(row, 3).value or '').strip().strip("'\"")
            attendings.append(_parse_attending_row(raw_name, prefs_raw, abs_raw))

        def names_in_col(col):
            names = set()
            for r in range(3, ws2.max_row + 1):
                v = ws2.cell(r, col).value
                if v:
                    s = str(v).strip().strip("'\"")
                    if s and not re.match(r'^\d+$', s) and s != 'None':
                        names.add(_lookup(s))
            return names

        last_year  = names_in_col(6)   # 24-25
        two_yr_ago = names_in_col(4)   # 23-24

        holiday_data = {
            'last_year':          sorted(last_year),
            'two_years_ago':      sorted(two_yr_ago),
            'holidays_two_years': sorted(last_year | two_yr_ago),
        }
        return attendings, holiday_data

# ─── Scheduling algorithm (faithful port of test_call_sched.py) ───────────────

def _overflow_assign(schedule, cd, atts, ro, is_hol, h2y_set):
    """Return the least-burdened eligible attending for an overflow week.
    Full-time attendings are preferred; part-time only used if no FT is eligible."""
    ft_candidates = []
    pt_candidates = []
    for name, status, _prefs, _abs in atts:
        if is_hol and name in h2y_set:
            continue
        if name in ro:
            continue
        if name in schedule.get(cd - datetime.timedelta(7), ""):
            continue
        if name in schedule.get(cd - datetime.timedelta(14), ""):
            continue
        if name in schedule.get(cd - datetime.timedelta(21), ""):
            continue
        count = _count(schedule, name)
        if status == 'full time':
            ft_candidates.append((name, count))
        else:
            pt_candidates.append((name, count))
    if ft_candidates:
        return min(ft_candidates, key=lambda x: x[1])[0]
    if pt_candidates:
        return min(pt_candidates, key=lambda x: x[1])[0]
    return None

def _week_in_range(start, end, date_ranges):
    cur = start
    while cur <= end:
        for dr in date_ranges:
            parts = dr.split('-')
            try:
                if len(parts) == 1:
                    d = datetime.datetime.strptime(dr.strip(), '%m/%d/%y').date()
                    if d == cur:
                        return True
                elif len(parts) == 2:
                    d1 = datetime.datetime.strptime(parts[0].strip(), '%m/%d/%y').date()
                    d2 = datetime.datetime.strptime(parts[1].strip(), '%m/%d/%y').date()
                    if d1 <= cur <= d2:
                        return True
            except ValueError:
                pass
        cur += datetime.timedelta(1)
    return False

def _count(schedule, name):
    return sum(1 for v in schedule.values() if name in str(v))

def run_schedule(attendings_data, holidays_cfg, h2y_set, start_str, end_str, max_retries=60):
    start = datetime.date.fromisoformat(start_str)
    end   = datetime.date.fromisoformat(end_str)
    total_weeks = (end - start).days // 7

    holidays = {k: datetime.date.fromisoformat(v) for k, v in holidays_cfg.items() if v}

    best_sched, best_unfilled = {}, float('inf')

    for _attempt in range(max_retries):
        atts = [(a['name'], a['status'], a.get('preferences', []), a.get('absences', []))
                for a in attendings_data]
        random.shuffle(atts)

        pt  = sum(1 for a in atts if a[1] in ('part time', 'part time 1'))
        ft  = len(atts) - pt
        ft_weeks  = total_weeks // ft if ft else 0
        pt_weeks  = ft_weeks - 1
        pt1_weeks = 1

        def quota(status):
            if status == 'full time':   return ft_weeks
            if status == 'part time':   return pt_weeks
            if status == 'part time 1': return pt1_weeks
            return ft_weeks

        # pre-build per-week lists
        req_off = {w: [] for w in range(total_weeks)}
        pref_wk = {w: [] for w in range(total_weeks)}
        cd = start
        for w in range(total_weeks):
            we = cd + datetime.timedelta(6)
            for a in atts:
                if _week_in_range(cd, we, a[3]): req_off[w].append(a[0])
                if _week_in_range(cd, we, a[2]): pref_wk[w].append(a[0])
            cd += datetime.timedelta(7)

        schedule = {}
        att_ctr = 0
        unfilled = 0
        cd = start

        for w in range(total_weeks):
            is_hol = any(
                (cd + datetime.timedelta(i)) == hd
                for hd in holidays.values()
                for i in range(7)
            )

            if att_ctr >= len(atts):
                att_ctr = 0

            ro   = req_off[w]
            pref = pref_wk[w]
            in_loop = True
            loop_cnt = 0
            safety   = 0
            finished_sched = 0  # reset each week: counts at-quota attendings seen this week

            while in_loop:
                safety += 1
                if safety > len(atts) * 10:
                    overflow = _overflow_assign(schedule, cd, atts, ro, is_hol, h2y_set)
                    if overflow:
                        schedule[cd] = f"{overflow} (extra)"
                    else:
                        schedule[cd] = "Need to hand schedule"
                        unfilled += 1
                    in_loop = False
                    break

                if att_ctr >= len(atts):
                    att_ctr = 0

                name, status = atts[att_ctr][0], atts[att_ctr][1]
                q = quota(status)
                appearances = _count(schedule, name)

                fourwk = schedule.get(cd - datetime.timedelta(28), "")
                fivewk = schedule.get(cd - datetime.timedelta(35), "")

                if pref and loop_cnt < len(atts):
                    loop_cnt += 1
                    can = (
                        appearances < q and
                        name in pref and
                        name not in schedule.get(cd - datetime.timedelta(7),  "") and
                        name not in schedule.get(cd - datetime.timedelta(14), "") and
                        name not in schedule.get(cd - datetime.timedelta(21), "") and
                        name not in fourwk and
                        name not in fivewk
                    )
                    if can:
                        schedule[cd] = f"{name} (preferred)"
                        att_ctr += 1
                        in_loop = False
                    else:
                        att_ctr += 1
                else:
                    hol_conflict = is_hol and (name in h2y_set)
                    prev_week = name in schedule.get(cd - datetime.timedelta(7), "")
                    if appearances < q and not hol_conflict and name not in ro and not prev_week:
                        schedule[cd] = name
                        att_ctr += 1
                        in_loop = False
                    else:
                        if appearances >= q:
                            finished_sched += 1
                            if finished_sched >= len(atts):
                                overflow = _overflow_assign(schedule, cd, atts, ro, is_hol, h2y_set)
                                if overflow:
                                    schedule[cd] = f"{overflow} (extra)"
                                else:
                                    schedule[cd] = "Need to hand schedule, all attendings have hit quota"
                                    unfilled += 1
                                in_loop = False
                        att_ctr += 1

            cd += datetime.timedelta(7)

        if unfilled < best_unfilled:
            best_unfilled = unfilled
            best_sched = dict(schedule)
        if best_unfilled == 0:
            break

    # Format output
    result = []
    cd = start
    for w in range(total_weeks):
        hol_names = [
            k.replace('_', ' ')
            for k, hd in holidays.items()
            if any((cd + datetime.timedelta(i)) == hd for i in range(7))
        ]
        result.append({
            'date':       cd.isoformat(),
            'week_end':   (cd + datetime.timedelta(6)).isoformat(),
            'assignment': best_sched.get(cd, 'Unassigned'),
            'is_holiday': bool(hol_names),
            'holidays':   hol_names,
        })
        cd += datetime.timedelta(7)

    return result, best_unfilled

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/load-excel')
def api_load_excel():
    if not os.path.exists(EXCEL_PATH):
        return jsonify({'error': f'Excel file not found: {EXCEL_PATH}'}), 404
    try:
        attendings, holiday_data = parse_excel(EXCEL_PATH)
        return jsonify({
            'attendings':    attendings,
            'holiday_data':  holiday_data,
            'holidays':      DEFAULT_HOLIDAYS,
            'start_date':    '2026-07-10',
            'end_date':      '2027-07-09',
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/upload-excel', methods=['POST'])
def api_upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xlsm')):
        return jsonify({'error': 'Please upload an .xlsx file'}), 400
    try:
        import io
        attendings, holiday_data = parse_excel(io.BytesIO(f.read()))
        return jsonify({
            'attendings':   attendings,
            'holiday_data': holiday_data,
            'holidays':     DEFAULT_HOLIDAYS,
            'start_date':   '2026-07-10',
            'end_date':     '2027-07-09',
            'filename':     f.filename,
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/generate', methods=['POST'])
def api_generate():
    data = request.json
    try:
        schedule, unfilled = run_schedule(
            attendings_data=data['attendings'],
            holidays_cfg=data.get('holidays', DEFAULT_HOLIDAYS),
            h2y_set=set(data.get('holidays_two_years', [])),
            start_str=data.get('start_date', '2026-07-10'),
            end_str=data.get('end_date', '2027-07-09'),
        )
        return jsonify({'schedule': schedule, 'unfilled': unfilled})
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/export-excel', methods=['POST'])
def api_export_excel():
    data = request.json
    schedule = data.get('schedule', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Schedule"

    header_fill   = PatternFill("solid", fgColor="2D3748")
    holiday_fill  = PatternFill("solid", fgColor="FFF3CD")
    preferred_fill = PatternFill("solid", fgColor="D1E7DD")
    unfilled_fill  = PatternFill("solid", fgColor="F8D7DA")
    header_font   = Font(bold=True, color="FFFFFF")

    for col, title in enumerate(["Week of", "Week ending", "Attending", "Notes"], 1):
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    def _fmt(iso):
        y, m, d = iso.split('-')
        return f"{int(m)}/{int(d)}/{y[2:]}"

    for entry in schedule:
        is_unfilled  = "Need to" in entry['assignment']
        is_preferred = "(preferred)" in entry['assignment']
        display_name = entry['assignment'].replace(" (preferred)", "")
        hol_notes    = ", ".join(h.replace("_", " ").title() for h in entry.get('holidays', []))
        r = ws.max_row + 1
        ws.cell(r, 1, _fmt(entry['date']))
        ws.cell(r, 2, _fmt(entry['week_end']))
        ws.cell(r, 3, display_name)
        ws.cell(r, 4, hol_notes)
        if is_unfilled:
            fill = unfilled_fill
        elif entry.get('is_holiday'):
            fill = holiday_fill
        elif is_preferred:
            fill = preferred_fill
        else:
            fill = None
        if fill:
            for c in range(1, 5):
                ws.cell(r, c).fill = fill

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='call_schedule.xlsx',
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
